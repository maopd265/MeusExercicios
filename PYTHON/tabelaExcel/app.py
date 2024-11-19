import openpyxl
#criar uma planilha (book)
book=openpyxl.Workbook()
#visualizar pag existente
print(book.sheetnames)
#criar pagina
book.create_sheet('Frutas')
#selecionar pagina
frutas_page=book['Frutas']
frutas_page.append(['FRUTAS','QUANTIDADE','PREÇO'])
frutas_page.append(['Teste','1','R$10'])
frutas_page.append(['Teste1','2','R$10'])
frutas_page.append(['Teste2','3','R$10'])
frutas_page.append(['Teste3','4','R$10'])
frutas_page.append(['Teste4','1','R$10'])
frutas_page.append(['Teste5','5','R$10'])
frutas_page.append(['Teste6','6','R$10'])
frutas_page.append(['Teste7','7','R$10'])
frutas_page.append(['Teste8','8','R$10'])

book.save('Planilhas de complas.xlsx')
